import openpyxl
from Sushi_class import Ingredients, SushiRoll
import re
import simplejson as json
import os

wb = openpyxl.load_workbook('sushi.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
maxr = sheet.max_row
maxc = sheet.max_column

ingredients_list = []

for rowNum in range(2, maxr):
    ingredient = Ingredients(sheet.cell(row=rowNum, column=1).value,
                sheet.cell(row=rowNum, column=2).value, sheet.cell(row=rowNum, column=3).value,
                sheet.cell(row=rowNum, column=4).value, sheet.cell(row=rowNum, column=5).value,
                sheet.cell(row=rowNum, column=6).value, sheet.cell(row=rowNum, column=7).value,
                sheet.cell(row=rowNum, column=8).value, int(rowNum))
    ingredients_list.append(ingredient)

if os.path.isfile("./sushi.json") and os.stat("./sushi.json").st_size != 0:
    old_file = open("./sushi.json", "r+")
    data = json.loads(old_file.read())
    # Opens existing list of sushi rolls
    rolls = []
    i = 0
    while i < 1:
        ingredients = []
        for item in data[i]["ingredients"]:
            for ingredient in ingredients_list:
                if item == ingredient.name:
                    ingredients_list.append(ingredient)
        roll = SushiRoll(data[i]["name"], data[i]["pieces"], data[i]["calories"],
                         data[i]["fat"], data[i]["protien"], data[i]["carbs"], ingredients)
        rolls.append(roll)
        i += 1
else:
    old_file = open("./sushi.json", "w+")
    rolls = []


def save():
    save_state = []
    for roll in rolls:
        roll_ingredients = []
        for ingredient in roll.ingredients:
            roll_ingredients.append(ingredient.name)
        roll_stats = {"name": roll.name,
                      "pieces": roll.get_pieces(),
                      "calories": roll.get_calories(),
                      "fat": roll.get_fat(),
                      "protien": roll.get_protien(),
                      "carbs": roll.get_carbs(),
                      "ingredients": roll_ingredients}
        save_state.append(roll_stats)
    old_file.seek(0)
    old_file.write(json.dumps(save_state))


def inputcheck(i):  # checks input and filters out non-numbers
    if i == "":
        return 1
    else:
        i = re.sub('[a-zA-z,.:()"=]', '', i)
        test = isinstance(int(i), int)
        if test is True:  # returns choice if valid, otherwise goes with option 1
            return i
        else:
            return 1


rice = []
condiments = []
veg_or_fruit = []
shell_fish = []
seaweed = []
roe = []
other_meat = []
ink_fish = []
fin_fish = []
meat_non_sea = []
eggs = []
wrappings = []
sushi_roll = []

for ingredient in ingredients_list:
    if ingredient.ingredient_type == "Rice":
        rice.append(ingredient)
    elif ingredient.ingredient_type == "Wrappings":
        wrappings.append(ingredient)
    elif ingredient.ingredient_type == "Eggs":
        eggs.append(ingredient)
    elif ingredient.ingredient_type == "Meat(None Seafood)":
        meat_non_sea.append(ingredient)
    elif ingredient.ingredient_type == "FinFish":
        fin_fish.append(ingredient)
    elif ingredient.ingredient_type == "Inkfish":
        ink_fish.append(ingredient)
    elif ingredient.ingredient_type == "Other Meat":
        other_meat.append(ingredient)
    elif ingredient.ingredient_type == "Roe":
        roe.append(ingredient)
    elif ingredient.ingredient_type == "Seaweed":
        seaweed.append(ingredient)
    elif ingredient.ingredient_type == "Shellfish":
        shell_fish.append(ingredient)
    elif ingredient.ingredient_type == "Veg/Fruit":
        veg_or_fruit.append(ingredient)
    else:
        condiments.append(ingredient)


def make_sushi_roll():
    print("Choose item type or enter 'Done':")
    i = 0
    type_list = []
    for ingredient in ingredients_list:
        if ingredient.ingredient_type not in type_list:
            print(i+1, ":" + str(ingredient.ingredient_type))
            type_list.append(ingredient.ingredient_type)
            i += 1
    choice = input()
    if choice == "Done":
        roll_name = input(print("Please enter roll name:"))
        roll_pieces = int(input(print("Please enter number of pieces in roll:")))
        fat, protien, carbs, calories = 0, 0, 0, 0
        for item in sushi_roll:
            fat += item.fat
            carbs += item.carbs
            protien += item.protien
            calories += item.calories
        roll_name = SushiRoll(str(roll_name), roll_pieces, calories, fat, protien, carbs, sushi_roll)
        rolls.append(roll_name)
        print("You created the ", roll_name.name + "\n" + "Nutrition facts: Calories:", roll_name.calories, "Fat:",
              roll_name.fat, "Protien:", roll_name.protien, "Carbs:", roll_name.carbs)
        save()
        start_screen()
    else:
        choice = int(inputcheck(choice))
        i = 0
        if choice == 1:
            for item in rice:
                print(item.idnum, str(item.name))
                i += 1
        if choice == 2:
            for item in wrappings:
                print(item.idnum, str(item.name))
                i += 1
        if choice == 3:
            for item in eggs:
                print(item.idnum, str(item.name))
                i += 1
        if choice == 4:
            for item in meat_non_sea:
                print(item.idnum, str(item.name))
                i += 1
        if choice == 5:
            for item in fin_fish:
                print(item.idnum, str(item.name))
                i += 1
        if choice == 6:
            for item in ink_fish:
                print(item.idnum, str(item.name))
                i += 1
        if choice == 7:
            for item in other_meat:
                print(item.idnum, str(item.name))
                i += 1
        if choice == 8:
            for item in roe:
                print(item.idnum, str(item.name))
                i += 1
        if choice == 9:
            for item in seaweed:
                print(item.idnum, str(item.name))
                i += 1
        if choice == 10:
            for item in shell_fish:
                print(item.idnum, str(item.name))
                i += 1
        if choice == 11:
            for item in veg_or_fruit:
                print(item.idnum, str(item.name))
                i += 1
        if choice == 12:
            for item in condiments:
                print(item.idnum, str(item.name))
                i += 1

        option = int(inputcheck(input())) - 2
        sushi_roll.append(ingredients_list[option])
        for item in sushi_roll:
            print(item.name)
        make_sushi_roll()


def start_screen():
    options = ["Make roll", "Adjust roll", "View rolls", "View Ingredients"]
    i = 1
    for option in options:
        print(i, ":", option)
        i += 1
    initial_choice = int(inputcheck(input()))
    if initial_choice == 1:
        make_sushi_roll()
    elif initial_choice == 2:
        roll_choice = view_rolls()
        adjust_roll(roll_choice)
    elif initial_choice == 3:
        x = view_rolls()
        i = 1
        for ingredient in rolls[x].ingredients:
            print(i, ":", ingredient.name, "-", "Calories:", ingredient.calories, "Fat:", ingredient.fat,
                  "Protien:", ingredient.protien, "Carbs:", ingredient.carbs)
            i += 1
        start_screen()
    else:
        i = 1
        for ingredient in ingredients_list:
            print(i, ":", ingredient.name)
            i += 1


def adjust_roll(x):
    roll_choice = x
    rolls[roll_choice].choose_action()
    roll_choice2 = int(inputcheck(input()))
    if roll_choice2 == 1:
        print("Choose ingredient to adjust:")
        i = 1
        for ingredient in rolls[roll_choice].ingredients:
            print(i, ":", ingredient.name)
            i += 1
        roll_choice3 = int(inputcheck(input())) -1
        print("Enter new amount (i.e. for 1/2 enter 0.5): ")
        adjustment = float(input())
        rolls[roll_choice].changing_servings(roll_choice3, adjustment)
        adjust_roll(roll_choice)
    elif roll_choice2 == 2:
        start_screen()
    elif roll_choice2 == 3:
        roll_choice = view_rolls()
        adjust_roll(roll_choice)
    else:
        start_screen()


def view_rolls():
    i = 1
    print("Choose roll:")
    for roll in rolls:
        print(i, ":", roll.name)
    roll_choice = int(inputcheck(input())) - 1
    return roll_choice

start_screen()
